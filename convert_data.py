import pandas as pd
import numpy as np
from unidecode import unidecode
import re
from tqdm import tqdm
import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials


def download_file_from_drive(docid, destination):
    scope = ['https://spreadsheets.google.com/feeds']
    credentials = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)
    client = gspread.authorize(credentials)
    spreadsheet = client.open_by_key(docid)
    wb = openpyxl.Workbook()
    for i, worksheet in enumerate(tqdm(spreadsheet.worksheets())):
        ws = wb.create_sheet(worksheet.title)
        data = worksheet.get_all_values()
        for dat in data:
            ws.append(dat)
    wb.remove(wb.get_sheet_by_name('Sheet'))
    wb.save(destination)


def download_raw_data():
    download_file_from_drive('1FfDe5f1SRRA_sEVENemxQmMjRi_bIrUQ9aLtN0XnPyY', 'raw_data/chatbot_data_core.xlsx')
    download_file_from_drive('1X_4Mnje7nrimj-at2erVeTkbHTqxczEwQ2VbVah1CFk', 'raw_data/chatbot_data_nlu.xlsx')


def convert_IE():
    data = pd.ExcelFile('raw_data/chatbot_data_nlu.xlsx')
    outfile = open('data/nlu.md', 'w', encoding='utf-8-sig')
    sheetnames = data.sheet_names[:-4]
    for name in sheetnames:
        sheet_data = data.parse(name)
        samples = sheet_data['User Request/Questions']
        intents = sheet_data['IntentName for Bot']
        for idx, sample in enumerate(samples):
            if len(intents) > idx and type(intents[idx]) == str:
                outfile.writelines('\n## intent:{}\n'.format(intents[idx]))
            if type(sample) == str:
                outfile.writelines(' - {}\n'.format(sample))
    outfile.close()


def convert_domain():
    df = pd.ExcelFile('raw_data/chatbot_data_nlu.xlsx')
    outfile = open('domain.yml', 'w', encoding='utf-8-sig')
    intent_dict = {}
    utter_dict = {}
    entities = []
    actions = []
    sheetnames = df.sheet_names[:-4]
    for name in sheetnames:
        sheet_data = df.parse(name)
        answers = sheet_data['Answer of Bot']
        intents = sheet_data['IntentName for Bot']
        questions = sheet_data['User Request/Questions']
        if 'Name of Utter' in sheet_data:
            utter_names = sheet_data['Name of Utter']
        else:
            print("Sheet {} don't have 'Name of Utter'".format(name))
        current_intent = 0
        for idx, answer in enumerate(answers):
            intent = intents[idx]
            if intent != current_intent and type(intent) == str:
                current_intent = intent
                intent_dict[intent] = answer
            if type(answer) == str:
                if 'Name of Utter' in sheet_data and type(utter_names[idx]) == str:
                    utter_dict[utter_names[idx]] = [answer.replace('"', "'").replace('\n', '\\n')]
                    actions.append(utter_names[idx])
                elif 'action_' not in answer:
                    if 'utter_{}'.format(current_intent) not in utter_dict:
                        utter_dict['utter_{}'.format(current_intent)] = []
                    utter_dict['utter_{}'.format(current_intent)].append(answer.replace('"', "'").replace('\n', '\\n'))
                    if 'utter_{}'.format(current_intent) not in actions:
                        actions.append('utter_{}'.format(current_intent))
                else:
                    if answer not in actions:
                        actions.append(answer)


        # find entities
        for question in questions:
            if type(question) == str:
                starts = [i+2 for i in range(len(question)) if question.startswith('](', i)]
                for start in starts:
                    end = question.find(')', start)
                    entity = question[start: end]
                    if entity not in entities:
                        entities.append(entity)
                    pass
    # read custom action
    custom_action_sheet = df.parse('Custom Action')
    for action in custom_action_sheet['Custom Action']:
        if action == action:
            actions.append(action)
    # read utter ask for required slot
    utter_ask = {}
    slots_sheet = df.parse('Slots')
    slots = slots_sheet['Slots']
    types = slots_sheet['Type']
    utters = slots_sheet['Utter_ask']
    set_actions = slots_sheet['Set Action']
    buttons = slots_sheet['Buttons']
    for idx, slot in enumerate(slots):
        if utters[idx] == utters[idx]:
            ask_str = utters[idx]
            if buttons[idx] == buttons[idx]:
                ask_str += '\n    buttons:\n'
                bts = buttons[idx].split(',')
                for bt in bts:
                    txt, payload = bt.split('/')
                    ask_str += '    - title: "{}"\n'.format(txt)
                    ask_str += '      payload: \'{}\'\n'.format(payload)

            utter_ask[slot] = ask_str

    # generate intents:
    outfile.writelines('intents:\n')
    rp_intents = []
    for intent in intent_dict:
        if '/' in intent:
            intent = intent.split('/')[0]
            if intent in rp_intents:
                continue
            rp_intents.append(intent)
            outfile.writelines('- {}:\n'.format(intent))
            outfile.writelines('    triggers: respond_{}\n'.format(intent))
        elif type(intent_dict[intent]) != str:
            outfile.writelines('- {}\n'.format(intent))
        elif 'action_' in intent_dict[intent]:
            outfile.writelines('- {}:\n'.format(intent))
            outfile.writelines('    triggers: {}\n'.format(intent_dict[intent]))
        else:
            outfile.writelines('- {}:\n'.format(intent))
            outfile.writelines('    triggers: utter_{}\n'.format(intent))
    # generate entities:
    outfile.writelines('entities:\n')
    for entity in entities:
        outfile.writelines('- {}\n'.format(entity))
    # generate slots:
    outfile.writelines('slots:\n')
    for idx, slot in enumerate(slots):
        outfile.writelines('  {}:\n'.format(slot))
        outfile.writelines('    type: {}\n'.format(types[idx]))
    outfile.writelines('  requested_slot:\n    type: text\n')
    # generate responses:
    outfile.writelines('responses:\n')
    outfile.writelines('  utter_default:\n  - text: "Xin lỗi mình không hiểu ý bạn ạ. Bạn nói rõ hơn được không!"\n')
    for utter in utter_dict:
        if '/' in utter:
            continue
        outfile.writelines('  {}:\n'.format(utter))
        for text in utter_dict[utter]:
            outfile.writelines('  - text: "{}"\n'.format(text))
    for slot in utter_ask:
        outfile.writelines('  utter_ask_{}:\n'.format(slot))
        outfile.writelines('  - text: {}\n'.format(utter_ask[slot]))
    # generate actions:
    outfile.writelines('actions:\n')
    outfile.writelines(' - utter_default\n')
    rp_actions = []
    for action in actions:
        if '/' in action:
            action = action.split('/')[0].replace('utter_', '')
            if action in rp_actions:
                continue
            rp_actions.append(action)
            outfile.writelines(' - respond_{}\n'.format(action))
        else:
            outfile.writelines(' - {}\n'.format(action))
    # generate set actions:
    for set_action in set_actions:
        if set_action == set_action:
            outfile.writelines(' - {}\n'.format(set_action))


def convert_stories():
    df = pd.ExcelFile('raw_data/chatbot_data_core.xlsx')
    outfile = open('data/stories.md', 'w')
    sheetnames = df.sheet_names
    for i, name in enumerate(sheetnames):
        sheet_data = df.parse(name)
        stts = sheet_data['STT']
        users = sheet_data['User']
        bots = sheet_data['Bot']
        stories = sheet_data['Rasa-Stories']
        for idx, action in enumerate(stories):
            if stts[idx] == stts[idx] and action == action:
                if idx != 0 or i != 0:
                    outfile.writelines('\n')
                story_name = '## {} {}\n'.format(name, int(stts[idx]))
                outfile.writelines(story_name)
            if users[idx] == users[idx] or '*' in action:
                action = action.replace('*', '').strip()
                outfile.writelines('* {}\n'.format(action.strip()))
            elif bots[idx] == bots[idx] or '-' in action:
                action = action.replace('-', '').strip()
                outfile.writelines('    - {}\n'.format(action.strip()))
    outfile.close()


def convert_respond():
    Frame = pd.ExcelFile('raw_data/chatbot_data_nlu.xlsx')
    outfile = open('data/responses.md', 'w', encoding='utf-8-sig')
    sheetnames = Frame.sheet_names[:-4]
    data = {}
    for name in sheetnames:
        if 'respond_' not in name.lower():
            continue
        sheet_data = Frame.parse(name)
        answers = sheet_data['Answer of Bot']
        intents = sheet_data['IntentName for Bot']
        curent_intent = 0
        for idx, answer in enumerate(answers):
            intent = intents[idx]
            if type(intent) == str:
                data[intent] = []
                curent_intent = intent
            if type(answer) == str:
                data[curent_intent].append(answer)
    for intent in data:
        outfile.writelines('\n## {}\n'.format(intent.split('/')[0]))
        outfile.writelines('* {}\n'.format(intent))
        answers = data[intent]
        for answer in answers:
            outfile.writelines('  - {}\n'.format(answer.replace('"', "'").replace('\n', '\\n')))

    outfile.close()


download_raw_data()
convert_IE()
convert_domain()
convert_stories()
convert_respond()